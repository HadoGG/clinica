from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import FileResponse
from django.db.models import Sum, Q, F
from django.utils import timezone
from datetime import datetime, timedelta
from django.contrib.auth.models import User
from django.views.decorators.cache import cache_page
from django.utils.decorators import method_decorator
from django.db import transaction

from payments.models import (
    Professional, Service, Attention, Discount, InsuranceDiscount,
    Settlement, SettlementLineItem, SettlementDiscount, UserProfile, AuditLog
)
from payments.serializers import (
    ProfessionalSerializer, ServiceSerializer, AttentionSerializer,
    DiscountSerializer, InsuranceDiscountSerializer, AuditLogSerializer,
    SettlementDetailSerializer, SettlementListSerializer,
    SettlementCreateSerializer, SettlementLineItemSerializer,
    SettlementDiscountSerializer, UserSerializer
)
from payments.permissions import IsAdmin, IsProfessional, IsAdminOrOwnAttention
from payments.audit import log_audit, get_changed_fields


class UserViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar usuarios (profesionales)"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    filterset_fields = ['profile__role']
    search_fields = ['username', 'first_name', 'last_name', 'email']
    ordering_fields = ['date_joined', 'first_name']
    ordering = ['-date_joined']
    
    def create(self, request, *args, **kwargs):
        """
        Crea un nuevo usuario con su perfil y roles asignados.
        
        Validaciones:
        - Username debe ser único
        - Email debe ser válido (opcional)
        - Contraseña: mínimo 10 caracteres + mayúscula + número + símbolo (del settings)
        
        Process:
        1. Valida los datos con UserSerializer
        2. Guarda usuario + contraseña hasheada + crea UserProfile (rol por defecto: 'professional')
        3. Si es profesional, crea automáticamente registro en Professional
        4. Refresca datos desde BD para obtener estado consistente
        5. Retorna respuesta con datos seguros (sin contraseña)
        
        Returns:
            201 Created: Usuario creado con todos los datos
            
        Raises:
            400 Bad Request: Si hay error en validación
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        # Refrescar el usuario desde la BD para obtener datos limpios
        user.refresh_from_db()
        
        # Devolver respuesta con datos seguros y serializables
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'is_active': user.is_active,
            'role': user.profile.role if hasattr(user, 'profile') else 'professional',
            'is_active_profile': user.profile.is_active if hasattr(user, 'profile') else user.is_active,
            'date_joined': user.date_joined.isoformat() if user.date_joined else None,
            'message': 'Usuario creado exitosamente'
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['get'])
    def professionals(self, request):
        """Obtiene solo usuarios con rol professional"""
        professionals = User.objects.filter(profile__role='professional')
        serializer = self.get_serializer(professionals, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Activa o desactiva un usuario"""
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        if hasattr(user, 'profile'):
            user.profile.is_active = user.is_active
            user.profile.save()
        return Response({
            'id': user.id,
            'username': user.username,
            'is_active': user.is_active,
            'message': f"Usuario {'activado' if user.is_active else 'desactivado'}"
        })
    
    @action(detail=True, methods=['post'])
    def change_role(self, request, pk=None):
        """Cambia el rol de un usuario"""
        user = self.get_object()
        new_role = request.data.get('role')
        
        if new_role not in ['admin', 'professional', 'staff']:
            return Response(
                {'error': 'Rol inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if hasattr(user, 'profile'):
            user.profile.role = new_role
            user.profile.save()
        else:
            UserProfile.objects.create(user=user, role=new_role)
        
        return Response({
            'id': user.id,
            'username': user.username,
            'role': new_role,
            'message': f"Rol cambiado a {new_role}"
        })
    
    @action(detail=True, methods=['post'])
    def set_password(self, request, pk=None):
        """Cambia la contraseña de un usuario"""
        user = self.get_object()
        password = request.data.get('password')
        
        if not password or len(password) < 6:
            return Response(
                {'error': 'La contraseña debe tener al menos 6 caracteres'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.set_password(password)
        user.save()
        
        return Response({
            'id': user.id,
            'username': user.username,
            'message': 'Contraseña actualizada exitosamente'
        })


class ProfessionalViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar profesionales"""
    queryset = Professional.objects.all()
    serializer_class = ProfessionalSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'specialization']
    search_fields = ['user__first_name', 'user__last_name', 'license_number']
    ordering_fields = ['created_at', 'user__first_name']
    ordering = ['-created_at']
    
    @action(detail=False, methods=['get'])
    def active_professionals(self, request):
        """Obtiene solo los profesionales activos, filtrado por rol"""
        user = request.user
        
        # Si es admin, puede ver todos los profesionales
        if hasattr(user, 'profile') and user.profile.role == 'admin':
            active = Professional.objects.filter(status='active')
        else:
            # Si no es admin, solo puede ver su propio perfil profesional
            try:
                active = Professional.objects.filter(user=user, status='active')
            except:
                active = Professional.objects.none()
        
        serializer = self.get_serializer(active, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def settlement_history(self, request, pk=None):
        """Obtiene el historial de liquidaciones de un profesional"""
        professional = self.get_object()
        settlements = professional.settlements.all()
        serializer = SettlementListSerializer(settlements, many=True)
        return Response(serializer.data)


class ServiceViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar servicios/prestaciones"""
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['is_active']
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'base_price']
    ordering = ['name']
    
    @method_decorator(cache_page(60 * 5))  # Cache 5 minutos si ENABLE_CACHE=True
    def list(self, request, *args, **kwargs):
        """Obtiene lista de servicios - Cacheado si se habilita"""
        return super().list(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'])
    def active_services(self, request):
        """Obtiene solo los servicios activos"""
        active = Service.objects.filter(is_active=True)
        serializer = self.get_serializer(active, many=True)
        return Response(serializer.data)


class AttentionViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar atenciones/procedimientos"""
    queryset = Attention.objects.all()
    serializer_class = AttentionSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['professional', 'service', 'status']
    search_fields = ['patient_name', 'patient_id']
    ordering_fields = ['date', 'amount_charged']
    ordering = ['-date']
    
    def get_queryset(self):
        """Filtrar atenciones según el rol del usuario"""
        user = self.request.user
        
        # Si es admin, puede ver todas las atenciones
        if hasattr(user, 'profile') and user.profile.role == 'admin':
            return Attention.objects.all()
        
        # Si no es admin, solo puede ver sus propias atenciones
        try:
            professional = user.professional_profile
            return Attention.objects.filter(professional=professional)
        except:
            # Si no tiene perfil de profesional, no puede ver ninguna atención
            return Attention.objects.none()
    
    @action(detail=False, methods=['get'])
    def professional_attentions(self, request):
        """Obtiene atenciones de un profesional específico"""
        professional_id = request.query_params.get('professional_id')
        if professional_id:
            attentions = Attention.objects.filter(professional_id=professional_id)
            serializer = self.get_serializer(attentions, many=True)
            return Response(serializer.data)
        return Response({'error': 'professional_id is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def date_range(self, request):
        """Obtiene atenciones en un rango de fechas"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date and end_date:
            try:
                start = datetime.fromisoformat(start_date)
                end = datetime.fromisoformat(end_date)
                attentions = Attention.objects.filter(date__range=[start, end])
                serializer = self.get_serializer(attentions, many=True)
                return Response(serializer.data)
            except ValueError:
                return Response({'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'error': 'start_date and end_date are required'}, status=status.HTTP_400_BAD_REQUEST)


class DiscountViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar descuentos y retenciones"""
    queryset = Discount.objects.all()
    serializer_class = DiscountSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['category', 'discount_type', 'is_active']
    search_fields = ['name']
    ordering_fields = ['name', 'value']
    ordering = ['name']
    
    @action(detail=False, methods=['get'])
    def active_discounts(self, request):
        """Obtiene solo los descuentos activos"""
        active = Discount.objects.filter(is_active=True)
        serializer = self.get_serializer(active, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_category(self, request):
        """Obtiene descuentos por categoría"""
        category = request.query_params.get('category')
        if category:
            discounts = Discount.objects.filter(category=category)
            serializer = self.get_serializer(discounts, many=True)
            return Response(serializer.data)
        return Response({'error': 'category is required'}, status=status.HTTP_400_BAD_REQUEST)


class SettlementViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar liquidaciones"""
    queryset = Settlement.objects.all()
    permission_classes = [IsAuthenticated]
    filterset_fields = ['professional', 'status']
    search_fields = ['professional__user__first_name', 'professional__user__last_name']
    ordering_fields = ['period_end', 'net_amount']
    ordering = ['-period_end']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return SettlementCreateSerializer
        elif self.action == 'retrieve':
            return SettlementDetailSerializer
        else:
            return SettlementListSerializer
    
    @action(detail=True, methods=['post'])
    def calculate(self, request, pk=None):
        """Calcula o recalcula una liquidación"""
        settlement = self.get_object()
        
        # Permitir cálculo en draft y calculated, pero no en approved o paid
        if settlement.status not in ['draft', 'calculated']:
            return Response(
                {'error': 'Solo se pueden calcular liquidaciones en estado borrador o calculado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            self._calculate_settlement(settlement)
            settlement.status = 'calculated'
            settlement.save()
            
            serializer = SettlementDetailSerializer(settlement)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Aprueba una liquidación calculada"""
        settlement = self.get_object()
        
        if settlement.status != 'calculated':
            return Response(
                {'error': 'Solo se pueden aprobar liquidaciones calculadas'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        settlement.status = 'approved'
        settlement.save()
        
        serializer = SettlementDetailSerializer(settlement)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """Exporta una liquidación como PDF"""
        settlement = self.get_object()
        
        try:
            from io import BytesIO
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, title=f"Liquidacion_{settlement.id}")
            
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#333333'),
                spaceAfter=30,
                alignment=1
            )
            story.append(Paragraph(f"Liquidación #{settlement.id}", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Información básica
            professional_name = settlement.professional.user.get_full_name() or settlement.professional.user.username
            info_data = [
                ['Profesional:', professional_name],
                ['Período:', f"{settlement.period_start.strftime('%d/%m/%Y')} - {settlement.period_end.strftime('%d/%m/%Y')}"],
                ['Estado:', settlement.status.upper()],
                ['Fecha de Generación:', settlement.created_at.strftime('%d/%m/%Y %H:%M')],
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#333333')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Tabla de montos
            amounts_data = [
                ['Concepto', 'Monto'],
                ['Comisión Total', f"${settlement.total_commission:,.2f}"],
                ['Descuentos', f"${settlement.total_discounts or 0:,.2f}"],
                ['Monto Neto', f"${settlement.net_amount:,.2f}"],
            ]
            
            amounts_table = Table(amounts_data, colWidths=[3*inch, 3*inch])
            amounts_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5b6ef5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(amounts_table)
            
            # Generar PDF
            doc.build(story)
            buffer.seek(0)
            
            return FileResponse(
                buffer,
                as_attachment=True,
                filename=f"liquidacion_{settlement.id}.pdf",
                content_type='application/pdf'
            )
        except Exception as e:
            return Response({
                'error': f'Error generando PDF: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporta liquidaciones como Excel"""
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Obtener todas las liquidaciones
            settlements = Settlement.objects.all().order_by('-period_end')
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Liquidaciones"
            
            # Estilos
            header_fill = PatternFill(start_color="5B6EF5", end_color="5B6EF5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['ID', 'Profesional', 'Período Inicio', 'Período Fin', 'Comisión Total', 'Descuentos', 'Monto Neto', 'Estado', 'Fecha Creación']
            ws.append(headers)
            
            # Aplicar estilos a headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Ancho de columnas
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 15
            
            # Datos
            for settlement in settlements:
                row = [
                    str(settlement.id),
                    settlement.professional.user.get_full_name() or settlement.professional.user.username,
                    settlement.period_start.strftime('%d/%m/%Y'),
                    settlement.period_end.strftime('%d/%m/%Y'),
                    settlement.total_commission,
                    settlement.total_discounts or 0,
                    settlement.net_amount,
                    settlement.status.upper(),
                    settlement.created_at.strftime('%d/%m/%Y %H:%M')
                ]
                ws.append(row)
                
                # Aplicar bordes a datos
                for cell in ws[ws.max_row]:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right' if cell.column > 1 else 'left')
            
            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return FileResponse(
                buffer,
                as_attachment=True,
                filename=f"liquidaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return Response({
                'error': f'Error generando Excel: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def mark_as_paid(self, request, pk=None):
        """Marca una liquidación como pagada"""
        settlement = self.get_object()
        
        if settlement.status != 'approved':
            return Response(
                {'error': 'Solo se pueden marcar como pagadas las liquidaciones aprobadas'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        payment_reference = request.data.get('payment_reference', '')
        
        settlement.status = 'paid'
        settlement.payment_date = timezone.now()
        settlement.payment_reference = payment_reference
        settlement.save()
        
        serializer = SettlementDetailSerializer(settlement)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def generate_for_period(self, request):
        """Genera liquidaciones para un período específico para todos los profesionales"""
        period_start = request.data.get('period_start')
        period_end = request.data.get('period_end')
        
        if not period_start or not period_end:
            return Response(
                {'error': 'period_start and period_end are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            start_date = datetime.fromisoformat(period_start).date()
            end_date = datetime.fromisoformat(period_end).date()
        except ValueError:
            return Response(
                {'error': 'Invalid date format'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        professionals = Professional.objects.filter(status='active')
        settlements = []
        
        for professional in professionals:
            settlement, created = Settlement.objects.get_or_create(
                professional=professional,
                period_start=start_date,
                period_end=end_date,
                defaults={'status': 'draft'}
            )
            settlements.append(settlement)
        
        serializer = SettlementListSerializer(settlements, many=True)
        return Response({
            'created_count': len([s for s in settlements if s.status == 'draft']),
            'settlements': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def report(self, request):
        """Genera un reporte de liquidaciones"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        status_filter = request.query_params.get('status')
        
        filters = Q()
        
        if start_date:
            try:
                start = datetime.fromisoformat(start_date).date()
                filters &= Q(period_start__gte=start)
            except ValueError:
                pass
        
        if end_date:
            try:
                end = datetime.fromisoformat(end_date).date()
                filters &= Q(period_end__lte=end)
            except ValueError:
                pass
        
        if status_filter:
            filters &= Q(status=status_filter)
        
        settlements = Settlement.objects.filter(filters)
        
        report_data = {
            'total_settlements': settlements.count(),
            'total_amount': float(settlements.aggregate(Sum('net_amount'))['net_amount__sum'] or 0),
            'by_status': {},
            'settlements': SettlementListSerializer(settlements, many=True).data
        }
        
        for status_choice in Settlement.STATUS_CHOICES:
            status_key = status_choice[0]
            count = settlements.filter(status=status_key).count()
            total = float(settlements.filter(status=status_key).aggregate(Sum('net_amount'))['net_amount__sum'] or 0)
            report_data['by_status'][status_choice[1]] = {
                'count': count,
                'total': total
            }
        
        return Response(report_data)
    
    @transaction.atomic  # ACID: Si falla algo, TODO se revierte
    def _calculate_settlement(self, settlement):
        """
        Calcula los montos de una liquidación (con garantía ACID).
        
        Descripción Detallada:
        1. Obtiene todas las atenciones completadas en el período de la liquidación
        2. Itera cada atención y calcula su comisión (considerando descuentos por aseguradora)
        3. Crea SettlementLineItem para cada atención (auditoría histórica)
        4. Aplica descuentos y retenciones obligatorias del sistema
        5. Registra cada descuento en SettlementDiscount (para trazabilidad)
        6. Calcula el monto neto final
        
        Garantía ACID (Atomicidad):
        Si falla en cualquier punto durante la ejecución:
        - Validación de datos
        - Cálculo de comisiones
        - Creación de ítems o descuentos
        - Guardado de liquidación
        TODAS las operaciones se revierten (rollback automático)
        No queda estado inconsistente en la BD
        
        Args:
            settlement (Settlement): La liquidación a calcular
            
        Raises:
            Exception: Si hay error (se revierte todo automáticamente)
            
        Side Effects:
        - Limpia SettlementLineItem previos de esta liquidación
        - Limpia SettlementDiscount previos de esta liquidación
        - Crea nuevos registros de línea para cada atención
        - Crea nuevos registros de descuento para cada descuento aplicado
        - Actualiza todos los campos totales de la liquidación
        """
        # Obtener todas las atenciones en el período que ya fueron completadas
        attentions = Attention.objects.filter(
            professional=settlement.professional,
            date__date__gte=settlement.period_start,
            date__date__lte=settlement.period_end,
            status='completed'  # Solo atenciones completadas
        )
        
        # Limpiar ítems previos
        SettlementLineItem.objects.filter(settlement=settlement).delete()
        SettlementDiscount.objects.filter(settlement=settlement).delete()
        
        # Calcular comisiones
        total_attended = 0
        total_commission = 0
        
        for attention in attentions:
            amount_charged = float(attention.amount_charged)
            total_attended += amount_charged
            
            commission = float(attention.calculate_commission())
            total_commission += commission
            
            SettlementLineItem.objects.create(
                settlement=settlement,
                attention=attention,
                service_name=attention.service.name,
                service_code=attention.service.code,
                attendance_date=attention.date.date(),
                amount_charged=attention.amount_charged,
                commission_percentage=attention.commission_percentage or attention.service.commission_percentage,
                commission_amount=commission
            )
        
        settlement.total_attended = total_attended
        settlement.total_commission = total_commission
        
        # Aplicar descuentos y retenciones
        total_discounts = 0
        total_retentions = 0
        
        discounts = Discount.objects.filter(is_active=True)
        
        for discount in discounts:
            if discount.category == 'discount':
                if discount.discount_type == 'percentage':
                    discount_amount = (total_commission * discount.value) / 100
                else:
                    discount_amount = float(discount.value)
                
                total_discounts += discount_amount
                
                SettlementDiscount.objects.create(
                    settlement=settlement,
                    discount=discount,
                    discount_type=discount.discount_type,
                    discount_value=discount.value,
                    discount_amount=discount_amount
                )
            
            elif discount.category == 'retention':
                if discount.discount_type == 'percentage':
                    discount_amount = (total_commission * discount.value) / 100
                else:
                    discount_amount = float(discount.value)
                
                total_retentions += discount_amount
                
                SettlementDiscount.objects.create(
                    settlement=settlement,
                    discount=discount,
                    discount_type=discount.discount_type,
                    discount_value=discount.value,
                    discount_amount=discount_amount
                )
        
        settlement.total_discounts = total_discounts
        settlement.total_retentions = total_retentions
        settlement.net_amount = total_commission - total_discounts - total_retentions
        settlement.save()


class InsuranceDiscountViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar descuentos por aseguradora"""
    queryset = InsuranceDiscount.objects.all()
    serializer_class = InsuranceDiscountSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    filterset_fields = ['is_active']
    search_fields = ['insurance_name']
    ordering_fields = ['insurance_name', 'discount_value']
    ordering = ['insurance_name']
    
    @action(detail=False, methods=['get'])
    def active_discounts(self, request):
        """Obtiene solo los descuentos por aseguradora activos"""
        active = InsuranceDiscount.objects.filter(is_active=True)
        serializer = self.get_serializer(active, many=True)
        return Response(serializer.data)


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para consultar logs de auditoría (solo lectura)"""
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    filterset_fields = ['user', 'action', 'model_name']
    search_fields = ['object_id', 'object_description', 'user__username']
    ordering_fields = ['created_at', 'user', 'action', 'model_name']
    ordering = ['-created_at']
    
    @action(detail=False, methods=['get'])
    def by_model(self, request):
        """Obtiene logs filtrados por modelo"""
        model_name = request.query_params.get('model_name')
        if model_name:
            logs = AuditLog.objects.filter(model_name=model_name)
            serializer = self.get_serializer(logs, many=True)
            return Response(serializer.data)
        return Response({'error': 'model_name is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def by_user(self, request):
        """Obtiene logs de cambios realizados por un usuario específico"""
        user_id = request.query_params.get('user_id')
        if user_id:
            logs = AuditLog.objects.filter(user_id=user_id)
            serializer = self.get_serializer(logs, many=True)
            return Response(serializer.data)
        return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Resumen de auditoría"""
        total_logs = AuditLog.objects.count()
        by_action = {}
        by_model = {}
        
        for action_choice in AuditLog.ACTION_CHOICES:
            action_key = action_choice[0]
            count = AuditLog.objects.filter(action=action_key).count()
            by_action[action_choice[1]] = count
        
        for model_choice in AuditLog.MODEL_CHOICES:
            model_key = model_choice[0]
            count = AuditLog.objects.filter(model_name=model_key).count()
            by_model[model_choice[1]] = count
        
        return Response({
            'total_logs': total_logs,
            'by_action': by_action,
            'by_model': by_model
        })
        
        settlement.total_discounts = total_discounts
        settlement.total_retentions = total_retentions
        settlement.net_amount = total_commission - total_discounts - total_retentions
        settlement.save()
