"""
Tareas asincrónicas con Celery para escalabilidad
"""
from celery import shared_task
from django.core.mail import send_mail
from django.template.loader import render_to_string
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from django.core.files.base import ContentFile
from payments.models import Settlement, Attention, AuditLog
import logging

logger = logging.getLogger(__name__)


@shared_task
def export_settlements_to_excel(period_start, period_end):
    """
    Tarea asincrónica: Exportar liquidaciones a Excel
    No bloquea al usuario mientras se genera el archivo
    
    Args:
        period_start: Fecha inicio (YYYY-MM-DD)
        period_end: Fecha fin (YYYY-MM-DD)
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Obtener liquidaciones del período
        settlements = Settlement.objects.filter(
            period_start__gte=period_start,
            period_end__lte=period_end
        ).order_by('-period_end')
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Liquidaciones"
        
        # Estilos
        header_fill = PatternFill(start_color="5B6EF5", end_color="5B6EF5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['ID', 'Profesional', 'Período Inicio', 'Período Fin', 
                  'Comisión Total', 'Descuentos', 'Monto Neto', 'Estado']
        ws.append(headers)
        
        # Aplicar estilos a headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for settlement in settlements:
            row = [
                str(settlement.id),
                settlement.professional.user.get_full_name() or settlement.professional.user.username,
                settlement.period_start.strftime('%d/%m/%Y'),
                settlement.period_end.strftime('%d/%m/%Y'),
                float(settlement.total_commission),
                float(settlement.total_discounts or 0),
                float(settlement.net_amount),
                settlement.status.upper(),
            ]
            ws.append(row)
            
            for cell in ws[ws.max_row]:
                cell.border = border
        
        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Excel exportado: {settlements.count()} liquidaciones")
        return {
            'status': 'success',
            'count': settlements.count(),
            'filename': f"liquidaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    
    except Exception as e:
        logger.error(f"Error exportando Excel: {str(e)}")
        return {'status': 'error', 'message': str(e)}


@shared_task
def generate_daily_report():
    """
    Tarea programada: Generar reporte diario de atenciones
    Se ejecuta automáticamente cada día a las 6 PM
    """
    try:
        today = datetime.now().date()
        
        # Atenciones del día
        attentions = Attention.objects.filter(
            date__date=today,
            status='completed'
        )
        
        total_amount = sum(float(a.amount_charged) for a in attentions)
        total_commission = sum(float(a.calculate_commission()) for a in attentions)
        
        report = {
            'date': today.isoformat(),
            'total_attentions': attentions.count(),
            'total_charged': float(total_amount),
            'total_commission': float(total_commission),
        }
        
        logger.info(f"Reporte diario generado: {report}")
        return report
    
    except Exception as e:
        logger.error(f"Error generando reporte diario: {str(e)}")
        return {'status': 'error', 'message': str(e)}


@shared_task
def cleanup_old_logs(days=30):
    """
    Tarea programada: Limpiar logs antiguos
    Se ejecuta automáticamente cada domingo a las 2 AM
    
    Args:
        days: Eliminar logs más antiguos que X días
    """
    try:
        cutoff_date = datetime.now() - timedelta(days=days)
        deleted_count, _ = AuditLog.objects.filter(created_at__lt=cutoff_date).delete()
        
        logger.info(f"Logs limpios: {deleted_count} registros eliminados (más antiguos de {days} días)")
        return {
            'status': 'success',
            'deleted': deleted_count,
            'days': days
        }
    
    except Exception as e:
        logger.error(f"Error limpiando logs: {str(e)}")
        return {'status': 'error', 'message': str(e)}


@shared_task
def send_settlement_notification(settlement_id):
    """
    Tarea asincrónica: Enviar notificación por email sobre liquidación
    
    Args:
        settlement_id: ID de la liquidación
    """
    try:
        settlement = Settlement.objects.get(id=settlement_id)
        professional_email = settlement.professional.user.email
        
        if not professional_email:
            logger.warning(f"No email para profesional {settlement.professional.id}")
            return {'status': 'skipped', 'reason': 'No email'}
        
        # Preparar email
        subject = f"Liquidación #{settlement.id} - Período {settlement.period_start} a {settlement.period_end}"
        html_message = render_to_string('settlement_notification.html', {
            'settlement': settlement,
            'professional': settlement.professional
        })
        
        # Enviar
        send_mail(
            subject=subject,
            message='Ver en el portal OdontAll',
            html_message=html_message,
            from_email='noreply@odontall.com',
            recipient_list=[professional_email],
            fail_silently=False,
        )
        
        logger.info(f"Notificación enviada a {professional_email}")
        return {'status': 'success', 'email': professional_email}
    
    except Settlement.DoesNotExist:
        logger.error(f"Liquidación {settlement_id} no encontrada")
        return {'status': 'error', 'message': 'Settlement not found'}
    
    except Exception as e:
        logger.error(f"Error enviando email: {str(e)}")
        return {'status': 'error', 'message': str(e)}


@shared_task
def calculate_settlement_async(settlement_id):
    """
    Tarea asincrónica: Calcular liquidación sin bloquear la UI
    
    Args:
        settlement_id: ID de la liquidación
    """
    try:
        from payments.views import SettlementViewSet
        
        settlement = Settlement.objects.get(id=settlement_id)
        view = SettlementViewSet()
        view._calculate_settlement(settlement)
        settlement.status = 'calculated'
        settlement.save()
        
        logger.info(f"Liquidación {settlement_id} calculada")
        return {'status': 'success', 'settlement_id': str(settlement_id)}
    
    except Settlement.DoesNotExist:
        logger.error(f"Liquidación {settlement_id} no encontrada")
        return {'status': 'error', 'message': 'Settlement not found'}
    
    except Exception as e:
        logger.error(f"Error calculando liquidación: {str(e)}")
        return {'status': 'error', 'message': str(e)}
