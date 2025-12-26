"""
Utilidades para generar reportes en PDF y Excel para OdontAll
"""
from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_settlement_pdf(settlement, filename=None):
    """
    Genera un PDF con los detalles de una liquidación
    
    Args:
        settlement: Objeto Settlement
        filename: Nombre del archivo (opcional)
    
    Returns:
        BytesIO con el contenido del PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                           topMargin=0.5*inch, bottomMargin=0.5*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph("OdontAll", title_style))
    elements.append(Paragraph("Liquidación de Pago", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información general
    info_data = [
        ['Profesional:', settlement.professional.user.get_full_name()],
        ['Período:', f"{settlement.period_start.strftime('%d/%m/%Y')} - {settlement.period_end.strftime('%d/%m/%Y')}"],
        ['Licencia:', settlement.professional.license_number],
        ['Especialidad:', settlement.professional.specialization],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Detalles de atenciones
    elements.append(Paragraph("Detalle de Atenciones", styles['Heading3']))
    
    line_items_data = [['Fecha', 'Servicio', 'Paciente', 'Monto', 'Comisión %', 'Comisión $']]
    
    for item in settlement.line_items.all():
        line_items_data.append([
            item.attendance_date.strftime('%d/%m/%Y'),
            item.service_name[:20],
            item.settlement.professional.user.get_full_name()[:15],
            f"${item.amount_charged:,.2f}",
            f"{item.commission_percentage}%",
            f"${item.commission_amount:,.2f}"
        ])
    
    line_table = Table(line_items_data, colWidths=[1.2*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
    line_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Resumen de liquidación
    elements.append(Paragraph("Resumen de Liquidación", styles['Heading3']))
    
    summary_data = [
        ['Concepto', 'Monto'],
        ['Total Atenciones', f"${settlement.total_attended:,.2f}"],
        ['Total Comisión', f"${settlement.total_commission:,.2f}"],
        ['Descuentos', f"-${settlement.total_discounts:,.2f}"],
        ['Retenciones', f"-${settlement.total_retentions:,.2f}"],
        ['NETO A PAGAR', f"${settlement.net_amount:,.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dcfce7')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Descuentos y retenciones aplicados
    if settlement.discounts_applied.exists():
        elements.append(Paragraph("Descuentos y Retenciones", styles['Heading3']))
        
        discounts_data = [['Concepto', 'Tipo', 'Valor', 'Monto']]
        
        for discount in settlement.discounts_applied.all():
            discounts_data.append([
                discount.discount.name[:25],
                discount.discount_type,
                f"{discount.discount_value}{'%' if discount.discount_type == 'percentage' else '$'}",
                f"${discount.discount_amount:,.2f}"
            ])
        
        discounts_table = Table(discounts_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        discounts_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(discounts_table)
    
    elements.append(Spacer(1, 0.5*inch))
    
    # Pie de página
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} - Sistema OdontAll",
        footer_style
    ))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_settlements_excel(settlements, filename=None):
    """
    Genera un archivo Excel con resumen de múltiples liquidaciones
    
    Args:
        settlements: QuerySet de Settlement
        filename: Nombre del archivo (opcional)
    
    Returns:
        BytesIO con el contenido del Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"
    
    # Estilos
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    total_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezado
    headers = ['Profesional', 'Período Inicio', 'Período Fin', 'Total Atenciones', 
               'Total Comisión', 'Descuentos', 'Retenciones', 'Neto a Pagar', 'Estado', 'Fecha Pago']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, settlement in enumerate(settlements, 2):
        ws.cell(row=row, column=1).value = settlement.professional.user.get_full_name()
        ws.cell(row=row, column=2).value = settlement.period_start
        ws.cell(row=row, column=3).value = settlement.period_end
        ws.cell(row=row, column=4).value = float(settlement.total_attended)
        ws.cell(row=row, column=5).value = float(settlement.total_commission)
        ws.cell(row=row, column=6).value = float(settlement.total_discounts)
        ws.cell(row=row, column=7).value = float(settlement.total_retentions)
        ws.cell(row=row, column=8).value = float(settlement.net_amount)
        ws.cell(row=row, column=9).value = settlement.get_status_display()
        ws.cell(row=row, column=10).value = settlement.payment_date
        
        # Aplicar estilos
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            
            # Formato de moneda para columnas de dinero
            if col in [4, 5, 6, 7, 8]:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            
            # Fondo para total
            if settlement.status == 'paid':
                cell.fill = total_fill
                cell.font = total_font
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    
    # Guardar en BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
